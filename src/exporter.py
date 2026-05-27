import os
from datetime import datetime

import pandas as pd
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from src.mapper import EXPORTE_COLUMNS

_PNM_HEADER_RENAME = {
    "PNM_R": "Status",
    "PNM_S": "Dw SNR",
    "PNM_T": "PL Dw",
    "PNM_U": "Up SNR",
    "PNM_V": "PL Up",
    "PNM_W": "CMTS Up",
    "PNM_X": "CMTS",
    "PNM_Y": "US Alias",
}

_AXTRACT_HEADER_RENAME = {
    "AXTRACT_ONT_STATUS":   "ONT Status",
    "AXTRACT_TX_POWER":     "Last TX Power (dBm)",
    "AXTRACT_RX_POWER":     "Last RX Power (dBm)",
    "AXTRACT_RX_OLT_POWER": "Last RX OLT Power (dBm)",
    "AXTRACT_RANGING":      "Ranging (m)",
    "AXTRACT_SFP_TYPE":     "SFP Type",
    "AXTRACT_FTTX_TIME":    "Last Down Time",
    "AXTRACT_ALARM_CODE":   "Last Down Cause",
    "AXTRACT_CMTS":         "OLT",
    "AXTRACT_CMTS_UP":      "Line Card",
    "AXTRACT_ARPON":        "ARPON AXTRACT",
    "AXTRACT_SPLITTER":     "Splitter AXTRACT",
    "AXTRACT_NAP":          "NAP AXTRACT",
    "AXTRACT_PUERTO_NAP":   "PUERTO NAP AXTRACT",
}

_PNM_COLS = set(_PNM_HEADER_RENAME.values()) | set(_AXTRACT_HEADER_RENAME.values())

# Anchos específicos del VBA (por letra de columna Excel, basados en posición fija del schema EXPORTE)
_LETTER_WIDTHS = {
    "D": 8,
    "E": 9.43,
    "F": 7.86,
    "G": 13.14,
    "J": 8.43,
    "P": 22.86,
}

_SORT_COLS = ["ID_NODO", "ID_AMPLIFICADOR", "ID_TAP"]

_HEADER_FILL = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
_PNM_FILL = PatternFill(start_color="BFBFBF", end_color="BFBFBF", fill_type="solid")
_RED_FILL    = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
_GREEN_FILL  = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
_YELLOW_FILL = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

_PNM_CF_RULES = [
    ("PNM_R", [
        ('{L}2="operational"', _GREEN_FILL),
        ('{L}2="rangingAutoAdjComplete"', _YELLOW_FILL),
        ('AND({L}2<>"",{L}2<>"operational",{L}2<>"rangingAutoAdjComplete")', _RED_FILL),
    ]),
    ("PNM_S", [
        ('AND({L}2<>"",{L}2>37)', _GREEN_FILL),
        ('AND({L}2>=35,{L}2<=37)', _YELLOW_FILL),
        ('AND({L}2<>"",{L}2<35)', _RED_FILL),
    ]),
    ("PNM_T", [
        ('AND({L}2<>"",{L}2>=-10,{L}2<=12)', _GREEN_FILL),
        ('AND({L}2<>"",{L}2>=-15,{L}2<-10)', _YELLOW_FILL),
        ('AND({L}2<>"",OR({L}2<-15,{L}2>12))', _RED_FILL),
    ]),
    ("PNM_U", [
        ('AND({L}2<>"",{L}2>29)', _GREEN_FILL),
        ('AND({L}2>=27,{L}2<=29)', _YELLOW_FILL),
        ('AND({L}2<>"",{L}2<27)', _RED_FILL),
    ]),
    ("PNM_V", [
        ('AND({L}2<>"",{L}2>=38,{L}2<=47.9)', _GREEN_FILL),
        ('AND({L}2<>"",{L}2>=48,{L}2<=50.9)', _YELLOW_FILL),
        ('AND({L}2<>"",OR({L}2<38,{L}2>50.9))', _RED_FILL),
    ]),
]


def _col_letter_for(col_name: str) -> str:
    idx = EXPORTE_COLUMNS.index(col_name)
    return get_column_letter(idx + 1)


def _format_exporte_sheet(ws, df: pd.DataFrame) -> None:
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    last_row = len(df) + 1
    header_font = Font(bold=True)

    for ci, col_name in enumerate(df.columns, start=1):
        letter = get_column_letter(ci)
        header_cell = ws.cell(row=1, column=ci)
        header_cell.font = header_font
        header_cell.alignment = Alignment(horizontal="center")
        header_cell.fill = _PNM_FILL if col_name in _PNM_COLS else _HEADER_FILL

        # Ancho de columna
        if letter in _LETTER_WIDTHS:
            ws.column_dimensions[letter].width = _LETTER_WIDTHS[letter]
        else:
            max_len = max(
                (len(str(cell.value)) if cell.value is not None else 0)
                for cell in ws[letter]
            )
            ws.column_dimensions[letter].width = min(max_len + 2, 50)

    # Formato de fecha en columna B (FECHA_DE_APERTURA)
    b_idx = EXPORTE_COLUMNS.index("FECHA_DE_APERTURA") + 1
    date_fmt = "DD/MM/YYYY HH:MM"
    for row in ws.iter_rows(min_row=2, max_row=last_row, min_col=b_idx, max_col=b_idx):
        for cell in row:
            cell.number_format = date_fmt

    # Conditional formatting: columna K roja si NRO_TIQUETE_TT ≠ "" y ≠ 0
    k_letter = _col_letter_for("NRO_TIQUETE_TT")
    ws.conditional_formatting.add(
        f"{k_letter}2:{k_letter}{last_row}",
        FormulaRule(formula=[f'AND({k_letter}2<>"",{k_letter}2<>0)'], fill=_RED_FILL),
    )

    for col_name, rules in _PNM_CF_RULES:
        letter = _col_letter_for(col_name)
        cf_range = f"{letter}2:{letter}{last_row}"
        for formula_template, fill in rules:
            ws.conditional_formatting.add(
                cf_range,
                FormulaRule(formula=[formula_template.replace("{L}", letter)], fill=fill),
            )

    # Helper dinámico para columnas AXTRACT (no están en EXPORTE_COLUMNS)
    cols_list = list(df.columns)

    def _ax_letter(col: str):
        try:
            return get_column_letter(cols_list.index(col) + 1)
        except ValueError:
            return None

    # CF simples AXTRACT
    _axtract_simple_cf = [
        ("ONT Status", [
            ('{L}2="up"',   _GREEN_FILL),
            ('{L}2="down"', _RED_FILL),
        ]),
        ("Last TX Power (dBm)", [
            ('AND({L}2>=0.5,{L}2<=5)',             _GREEN_FILL),
            ('AND({L}2<>"",OR({L}2<0.5,{L}2>5))', _RED_FILL),
        ]),
        ("Last RX Power (dBm)", [
            ('AND({L}2>=-27,{L}2<=-10)',                  _GREEN_FILL),
            ('AND({L}2<>"",OR({L}2<-27,{L}2>-10))',      _RED_FILL),
        ]),
    ]
    for col_name, rules in _axtract_simple_cf:
        letter = _ax_letter(col_name)
        if not letter:
            continue
        cf_range = f"{letter}2:{letter}{last_row}"
        for formula_template, fill in rules:
            ws.conditional_formatting.add(
                cf_range,
                FormulaRule(formula=[formula_template.replace("{L}", letter)], fill=fill),
            )

    # CF dependientes de SFP Type (Last RX OLT Power y Ranging)
    sfp = _ax_letter("SFP Type")
    if sfp:
        olt_rx = _ax_letter("Last RX OLT Power (dBm)")
        if olt_rx:
            olt_range = f"{olt_rx}2:{olt_rx}{last_row}"
            for sfp_type, lo, hi in [("classbplus", -28, -10), ("classcplus", -32, -14), ("classcplusplus", -35, -17)]:
                ws.conditional_formatting.add(olt_range, FormulaRule(
                    formula=[f'AND({sfp}2="{sfp_type}",{olt_rx}2>={lo},{olt_rx}2<={hi})'], fill=_GREEN_FILL))
                ws.conditional_formatting.add(olt_range, FormulaRule(
                    formula=[f'AND({sfp}2="{sfp_type}",{olt_rx}2<>"",OR({olt_rx}2<{lo},{olt_rx}2>{hi}))'], fill=_RED_FILL))

        ran = _ax_letter("Ranging (m)")
        if ran:
            ran_range = f"{ran}2:{ran}{last_row}"
            for sfp_type, threshold in [("classbplus", 10000), ("classcplus", 15000), ("classcplusplus", 20000)]:
                ws.conditional_formatting.add(ran_range, FormulaRule(
                    formula=[f'AND({sfp}2="{sfp_type}",{ran}2<={threshold})'], fill=_GREEN_FILL))
                ws.conditional_formatting.add(ran_range, FormulaRule(
                    formula=[f'AND({sfp}2="{sfp_type}",{ran}2>{threshold})'], fill=_RED_FILL))


def export_to_excel(
    df_exporte: pd.DataFrame,
    df_raw: pd.DataFrame,
    df_axtract: pd.DataFrame,
    df_pnm: pd.DataFrame,
    output_dir: str,
    username: str = "",
) -> str:
    os.makedirs(output_dir, exist_ok=True)

    timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M")
    user_suffix = f"_{username}" if username else ""
    filename = f"Ingreso_Siebel_{timestamp}{user_suffix}.xlsx"
    filepath = os.path.join(output_dir, filename)

    sort_cols = [c for c in _SORT_COLS if c in df_exporte.columns]
    df_sorted = df_exporte.sort_values(sort_cols, na_position="last").reset_index(drop=True)
    df_excel = df_sorted.rename(columns={**_PNM_HEADER_RENAME, **_AXTRACT_HEADER_RENAME})

    with pd.ExcelWriter(filepath, engine="openpyxl") as writer:
        df_excel.to_excel(writer, sheet_name="EXPORTE", index=False)
        df_raw.to_excel(writer, sheet_name="RAW", index=False)

        _format_exporte_sheet(writer.sheets["EXPORTE"], df_excel)

        ws_raw = writer.sheets["RAW"]
        ws_raw.freeze_panes = "A2"
        for col_cells in ws_raw.columns:
            max_len = max(
                (len(str(c.value)) if c.value is not None else 0) for c in col_cells
            )
            ws_raw.column_dimensions[col_cells[0].column_letter].width = min(max_len + 2, 50)

        if not df_axtract.empty:
            df_axtract.to_excel(writer, sheet_name="AXTRACT", index=False)
            ws_ast = writer.sheets["AXTRACT"]
            ws_ast.freeze_panes = "A2"
            for col_cells in ws_ast.columns:
                max_len = max(
                    (len(str(c.value)) if c.value is not None else 0) for c in col_cells
                )
                ws_ast.column_dimensions[col_cells[0].column_letter].width = min(max_len + 2, 50)

        if not df_pnm.empty:
            df_pnm.to_excel(writer, sheet_name="PNM", index=False)
            ws_pnm = writer.sheets["PNM"]
            ws_pnm.freeze_panes = "A2"
            for col_cells in ws_pnm.columns:
                max_len = max(
                    (len(str(c.value)) if c.value is not None else 0) for c in col_cells
                )
                ws_pnm.column_dimensions[col_cells[0].column_letter].width = min(max_len + 2, 50)

    return os.path.abspath(filepath)
